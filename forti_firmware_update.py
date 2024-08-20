import paramiko
import openpyxl

def create_ssh_client(hostname, port, username, password):
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    try:
        ssh.connect(hostname, port, username, password)
        print("Conexión SSH establecida correctamente.")
    except paramiko.AuthenticationException:
        print("Error de autenticación. Verifica tu nombre de usuario y contraseña.")
        return None
    except paramiko.SSHException as sshException:
        print(f"Error al conectar a {hostname}: {sshException}")
        return None
    except Exception as e:
        print(f"Error al conectar a {hostname}: {e}")
        return None
    return ssh

def get_firmware_version(ssh):
    try:
        command = 'get system status'
        stdin, stdout, stderr = ssh.exec_command(command)
        stdout.channel.recv_exit_status()  # Esperar a que el comando termine
        
        output = stdout.read().decode()
        errors = stderr.read().decode()
        
        if errors:
            print(f"Errores: {errors}")
            
        for line in output.splitlines():
            if "Version:" in line:
                version = line.split("Version:")[1].strip().split(',')[0]
                print(f"Versión del firmware: {version}")
                return version
            
    except Exception as e:
        print(f"Error al obtener la versión del firmware: {e}")
        return None

def get_hostname(ssh):
    try:
        command = 'get system status'
        stdin, stdout, stderr = ssh.exec_command(command)
        stdout.channel.recv_exit_status()  # Esperar a que el comando termine
        
        output = stdout.read().decode()
        errors = stderr.read().decode()
        
        if errors:
            print(f"Errores: {errors}")
        
        for line in output.splitlines():
            if "Hostname:" in line:
                hostname = line.split("Hostname:")[1].strip()
                print(f'Hostname: {hostname}')
                return hostname
            
    except Exception as e:
        print(f"Error al obtener el hostname: {e}")
        return None

def save_to_excel(hostname, device_ip, username, password, version, file_name='Forti_firmware_version.xlsx'):
    try:
        # Cargar el archivo de Excel existente
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
        print(f"Archivo Excel '{file_name}' cargado correctamente.")

        # Encontrar la primera fila vacía
        row = sheet.max_row + 1
        
        # Asignar los valores a las columnas correspondientes  
        sheet.cell(row=row, column=2, value=hostname)
        sheet.cell(row=row, column=3, value=device_ip)
        sheet.cell(row=row, column=4, value=username)  
        sheet.cell(row=row, column=5, value=password)  
        sheet.cell(row=row, column=6, value=version)   

        # Guardar el archivo Excel
        workbook.save(file_name)
        print(f"Datos guardados en '{file_name}' en la fila {row}.")
        
    except Exception as e:
        print(f"Error al guardar los datos en Excel: {e}")

def main():
    # Configuración del Fortigate
    device_ip = '190.12.62.187'      # Dirección IP del Fortigate
    port = 22                       # Puerto SSH
    username = 'admin'              # Usuario SSH
    password = 'FW_sail$$2020'      # Contraseña SSH

    ssh = None
    try:
        # Crear una conexión SSH
        ssh = create_ssh_client(device_ip, port, username, password)
        
        if ssh:
            # Obtener el hostname y la versión del firmware
            hostname = get_hostname(ssh)
            version = get_firmware_version(ssh)

            if hostname and version:
                # Guardar los datos en el archivo Excel existente
                save_to_excel(hostname, device_ip, username, password, version)
                
    finally:
        if ssh:
            # Cerrar la conexión SSH
            ssh.close()

if __name__ == "__main__":
    main()
